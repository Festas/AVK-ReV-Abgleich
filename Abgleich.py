from openpyxl import load_workbook  
from pathlib import Path, PureWindowsPath
from time import sleep
print("erstellt von Gregor Schuboth")
sleep(1)
pfad=str(PureWindowsPath(r"Z:\KKK\Fachbereiche\TKW\Allgemein\ReVS Arbeitsordner\12_ReVS-AVK-Abgleich"))
revs=load_workbook(pfad+"/export.xlsx").active
if pfad+"/AVK.xlsx"==True:
    avk=load_workbook(pfad+"/AVK.xlsx").active
elif pfad+"/Avk.xlsx"==True:
    avk=load_workbook(pfad+"/Avk.xlsx").active
else:
    avk=load_workbook(pfad+"/avk.xlsx").active
Abgleich=avk  

spalteReVS=["",
        "Verpackungs-ID",
        "Standort",
        "Reststoff-ID",
        "Nettomasse/kg",
        "AVK-Übersetzungsstandort"] 
spalteAVK=["",
        "Behälternummer",
        "Lagerort/Absender",
        "Abfallmasse [kg]",
        "Individuelle ID"] #Position in AVK-Liste
aufgenommenReVS=[]
aufgenommenAVK=[]
Ort=0 #ändert sich, wenn im AVK-Auszug KKK als Ort berücksichtigt wird
#einlesen der gesuchten Spalten

h=[]
for i in range(1,avk.max_column+1):
        h.append(avk.cell(1,i).value)
for i in range(1,len(spalteAVK)):
    spalteAVK[i]=h.index(spalteAVK[i])+1  
h=[]
for i in range(1,revs.max_column+1):
    h.append(revs.cell(1,i).value)
Abgleich.active.delete_cols(1,Abgleich.active.max_column)
for i in range(1,len(spalteReVS)-1):
    spalteReVS[i]=h.index(spalteReVS[i])+1 

restrevs=[]
for i in range(len(spalteReVS)):
    restrevs.append("")
restavk=[]
for i in range(len(spalteAVK)):
    restavk.append("")

GebindeReVS=[]
GebindeReVS.append(spalteReVS.copy())
GebindeAVK=[]
GebindeAVK.append(spalteAVK.copy())

Fehler=[["Verpackungs-ID","Reststoff-ID","Individuelle ID","Standort ReVS","Standort AVK", "ReVS Nettomasse/kg", "AVK Abfallmasse [kg]"]]

for i in range(1,avk.max_row+1):      #AVK einlesen    
    GebindeAVK.append(restavk.copy())
    if avk.cell(i,spalteAVK[1]).value!="":
        GebindeAVK[i][0]=avk.cell(i,spalteAVK[1]).value[0:avk.cell(i,spalteAVK[1]).value.find(" ")]
        if avk.cell(i,spalteAVK[1]).value.count("*")!=0:
            GebindeAVK[i][1]=avk.cell(i,spalteAVK[1]).value[0:avk.cell(i,spalteAVK[1]).value.find("*")]
        elif avk.cell(i,spalteAVK[1]).value[0:3]=="KKK":
            GebindeAVK[i][0]=avk.cell(i,spalteAVK[1]).value[0:3]
            GebindeAVK[i][1]=avk.cell(i,spalteAVK[1]).value
        else:
            GebindeAVK[i][1]=avk.cell(i,spalteAVK[1]).value
        if aufgenommenAVK.count(GebindeAVK[i][0])==0 and GebindeAVK[i][0]!="":
            aufgenommenAVK.append(GebindeAVK[i][0])
        GebindeAVK[i][2]=avk.cell(i,spalteAVK[2]).value
        GebindeAVK[i][3]=avk.cell(i,spalteAVK[3]).value
        GebindeAVK[i][4]=avk.cell(i,spalteAVK[4]).value
        if GebindeAVK[i][2]=="KKK":
            Ort=1

for i in range(1,revs.max_row+1):  #ReVS einlesen
        GebindeReVS.append(restrevs.copy())
        if revs.cell(i,spalteReVS[1]).value!="":
            GebindeReVS[i][0]=revs.cell(i,spalteReVS[1]).value[0:revs.cell(i,spalteReVS[1]).value.find("-")] #Gebindetyp
        GebindeReVS[i][2]=revs.cell(i,spalteReVS[2]).value
        if aufgenommenAVK.count(GebindeReVS[i][0])!=0:      
            if GebindeReVS[i][0][0:1]=="V" or GebindeReVS[i][0][0:2]=="MH" or GebindeReVS[i][0][0:2]=="SO" or GebindeReVS[i][0][0:1]=="E" or GebindeReVS[i][0][0:2]=="AK" or GebindeReVS[i][0][0:2]=="KE": #Unterscheidung, da AVK keine einheitliche Bezeichnung hat
                GebindeReVS[i][1]=GebindeReVS[i][0]+" "+revs.cell(i,spalteReVS[1]).value[revs.cell(i,spalteReVS[1]).value.find("-")+2:revs.cell(i,spalteReVS[1]).value.find("-")+5] #dreistellige Nummern
            elif GebindeReVS[i][0][0:4]=="KKK":
                GebindeReVS[i][1]=GebindeReVS[i][1]
            else:
                GebindeReVS[i][1]=GebindeReVS[i][0]+" "+revs.cell(i,spalteReVS[1]).value[revs.cell(i,spalteReVS[1]).value.find("-")+1:revs.cell(i,spalteReVS[1]).value.find("-")+7] #vierstellige Nummern
            if aufgenommenReVS.count(GebindeReVS[i][0])==0 and GebindeReVS[i][0]!="":
                aufgenommenReVS.append(GebindeReVS[i][0])
            GebindeReVS[i][3]=revs.cell(i,spalteReVS[3]).value
            GebindeReVS[i][4]=revs.cell(i,spalteReVS[4]).value
            if GebindeReVS[i][2][0:2]=="EX":
                GebindeReVS[i][5]="KKK"
            elif GebindeReVS[i][2]=="ZW6-1":
                GebindeReVS[i][5]="W 06"
            elif  GebindeReVS[i][2][0:2]=="A-" or GebindeReVS[i][2][0:2]=="AL" or GebindeReVS[i][2][0:1]=="E" or GebindeReVS[i][2][0:2]=="ST":
                GebindeReVS[i][5]=revs.cell(i,spalteReVS[2]).value[revs.cell(i,spalteReVS[2]).value.find("Z")+1:revs.cell(i,spalteReVS[2]).value.find("Z")+2]+ " "+revs.cell(i,spalteReVS[2]).value[revs.cell(i,spalteReVS[2]).value.find("Z")+2:revs.cell(i,spalteReVS[2]).value.find("Z")+7]  
            elif GebindeReVS[i][2]=="Containerstellplatz-ÜB":
                GebindeReVS[i][5]='CONTAINER 20"'
            for j in range(1,len(GebindeAVK)):
                if GebindeAVK[j][0]!="vorhanden": #überspringen gleicher Gebinde
                    if GebindeReVS[i][1]==GebindeAVK[j][1]:
                        GebindeReVS[i][0]="vorhanden"
                        GebindeAVK[j][0]="vorhanden"
                        Fehler.append(Fehler[0].copy())
                        x=0
                        Fehler[-1][0]=GebindeReVS[i][1]
                        Fehler[-1][1]=GebindeReVS[i][3]
                        Fehler[-1][2]=GebindeAVK[j][4]
                        Fehler[-1][3]=GebindeReVS[i][2]
                        Fehler[-1][4]=GebindeAVK[j][2]
                        Fehler[-1][5]=GebindeReVS[i][4]
                        Fehler[-1][6]=GebindeAVK[j][3]
                        if Fehler[-1][1]==Fehler[-1][2]:
                            Fehler[-1][2]="stimmt mit ReVS"
                        else:
                            x+=1
                        if GebindeReVS[i][5]==GebindeAVK[j][2]:
                            Fehler[-1][4]="stimmt mit ReVS"
                        else:
                            x+=1
                        if Fehler[-1][5]==Fehler[-1][6]:
                            Fehler[-1][6]="stimmt mit ReVS"
                        else:
                            x+=1
                        if x==0:
                            del Fehler[-1]                               
            if aufgenommenAVK.count(GebindeReVS[i][0])!=0 and (GebindeReVS[i][5]!="KKK" and Ort==0) and GebindeReVS[i][2]!="An AVK übergeben" :
                Fehler.append(Fehler[0].copy())
                Fehler[x][0]=GebindeReVS[i][1]
                Fehler[x][1]=GebindeReVS[i][3]
                Fehler[x][2]=""
                Fehler[x][3]=GebindeReVS[i][2]
                Fehler[x][4]="nicht im AVK gelistet"
                Fehler[x][5]=GebindeReVS[i][4]
                Fehler[x][6]=""

        
                
for i in range(1,len(GebindeAVK)):  #geht alle Gebinde im AVK durch
    if  GebindeAVK[i][0]!="vorhanden" and aufgenommenReVS.count(GebindeAVK[i][0])!=0:      #überspringen gleicher Gebinde
        x=x+1
        Fehler.append(Fehler[0].copy())
        Fehler[x][0]=GebindeAVK[i][1]
        Fehler[x][1]="nicht im ReVS gelistet"
        Fehler[x][2]=GebindeAVK[i][4]
        Fehler[x][3]=""
        Fehler[x][4]=GebindeAVK[i][2]
        Fehler[x][5]=""
        Fehler[x][6]=GebindeAVK[i][3]
Abgleich.active.cell(1,1).value="Fehlerhafte und fehlende Gebinde"
for i in range(1,len(Fehler)+1):
    for j in range(1,len(Fehler[i-1])+1):
        Abgleich.active.cell(i+1,j).value=Fehler[i-1][j-1]
Abgleich.save(pfad+"/Abgleich.xlsx")
#Abgleich.save("Desktop/Abgleich.xlsx")
Abgleich.close()
